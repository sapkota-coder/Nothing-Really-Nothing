import tkinter as tk
from tkinter import ttk, messagebox
import os
import time
import threading
import json
from datetime import datetime, timedelta
import openpyxl

DATA_FILE = "Data/input_data.json"
EXPORT_DIR = "Data"

class CardFrame(ttk.Frame):
    """Custom Frame with padding and border for card effect."""
    def __init__(self, parent, title="", **kwargs):
        super().__init__(parent, padding=15, style="Card.TFrame", **kwargs)
        if title:
            self.title_label = ttk.Label(self, text=title, style="CardTitle.TLabel")
            self.title_label.pack(anchor="w", pady=(0, 10))

class FinancialApp:
    def __init__(self, root):
        self.root = root
        self.root.title("E-Commerce Profit Calculator")
        self.entries = {}
        self.fields_order = []
        self.result_labels = {}
        self.current_results = {}
        self.animating = {}

        self.build_style()
        self.build_ui()
        self.load_data()
        self.update_results()
        self.schedule_midnight_save()

    def build_style(self):
        style = ttk.Style(self.root)
        self.root.configure(bg="#121212")
        style.theme_use("clam")

        # Card frame style
        style.configure("Card.TFrame", background="#1F1F1F", relief="ridge", borderwidth=1)
        style.configure("CardTitle.TLabel", font=("Segoe UI", 14, "bold"), foreground="#00FFF7", background="#1F1F1F")

        # Labels
        style.configure("TLabel", background="#121212", foreground="cyan", font=("Segoe UI", 11))
        style.configure("InputLabel.TLabel", font=("Segoe UI", 11), foreground="white", background="#1F1F1F")
        style.configure("ResultLabel.TLabel", font=("Segoe UI", 11), foreground="#B0F0EF", background="#1F1F1F")
        style.configure("ResultValue.TLabel", font=("Segoe UI", 12, "bold"), foreground="#00FFF7", background="#1F1F1F")
        style.configure("BigResult.TLabel", font=("Segoe UI", 16, "bold"), foreground="#00FFF7", background="#121212")

        # Entries
        style.configure("TEntry",
                        fieldbackground="#292929",
                        foreground="cyan",
                        bordercolor="#333333",
                        lightcolor="#00FFF7",
                        darkcolor="#121212",
                        borderwidth=1,
                        relief="flat",
                        padding=6,
                        font=("Segoe UI", 12))
        style.map("TEntry",
                  focusbackground=[("focus", "#00FFF7")],
                  foreground=[("focus", "#00FFF7")])

        # Button
        style.configure("TButton",
                        background="#00FFF7",
                        foreground="#121212",
                        font=("Segoe UI", 12, "bold"),
                        padding=10)
        style.map("TButton",
                  background=[("active", "#00D6CC")])

    def build_ui(self):
        self.root.geometry("900x700")
        self.root.minsize(850, 650)

        main_container = ttk.Frame(self.root, padding=20, style="TLabel")
        main_container.pack(fill="both", expand=True)

        title = ttk.Label(main_container, text="E-Commerce Profit Calculator",
                          font=("Segoe UI", 24, "bold"), foreground="#00FFF7", background="#121212")
        title.pack(anchor="center", pady=(0, 20))

        canvas = tk.Canvas(main_container, bg="#121212", highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, style="TLabel")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Enable mouse wheel scrolling on Windows
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Input card
        input_card = CardFrame(scrollable_frame, "Input Parameters")
        input_card.pack(side="left", fill="both", expand=True, padx=(0, 15), pady=10)

        # Result card
        result_card = CardFrame(scrollable_frame, "Calculation Results")
        result_card.pack(side="left", fill="both", expand=True, pady=10)

        # Create internal frame inside result_card for grid usage
        result_grid_frame = ttk.Frame(result_card, style="Card.TFrame")
        result_grid_frame.pack(fill="both", expand=True)

        # Group inputs logically
        group1 = ttk.LabelFrame(input_card, text="Revenue & Costs", padding=10, style="Card.TFrame")
        group1.pack(fill="x", pady=(0, 15))

        group1_fields = ["Sales", "COGS", "Shipping", "Warehousing", "Platform Commissions"]
        self.create_input_fields(group1, group1_fields)

        group2 = ttk.LabelFrame(input_card, text="Marketing & Operations", padding=10, style="Card.TFrame")
        group2.pack(fill="x", pady=(0, 15))

        group2_fields = ["Marketing", "Technology", "Fixed Cost", "Depreciation", "Interest", "Tax Rate (%)"]
        self.create_input_fields(group2, group2_fields)

        group3 = ttk.LabelFrame(input_card, text="Ads & Customers", padding=10, style="Card.TFrame")
        group3.pack(fill="x", pady=(0, 0))

        group3_fields = ["Ad Spend", "Customers"]
        self.create_input_fields(group3, group3_fields)

        # Configure grid in result_grid_frame
        result_grid_frame.grid_columnconfigure(0, weight=1)
        result_grid_frame.grid_columnconfigure(1, weight=1)

        intermediate_results = [
            "Gross Margin", "Contribution Margin 1", "Contribution Margin 2",
            "EBITDA", "EBIT", "PBT", "TAX", "PAT", "Customer Acquisition Cost"
        ]

        for idx, field in enumerate(intermediate_results):
            row = idx // 2
            col = (idx % 2) * 2

            lbl_name = ttk.Label(result_grid_frame, text=field + ":", style="ResultLabel.TLabel")
            lbl_name.grid(row=row, column=col, sticky="e", padx=10, pady=8)

            lbl_val = ttk.Label(result_grid_frame, text="0.00", style="ResultValue.TLabel")
            lbl_val.grid(row=row, column=col + 1, sticky="w", padx=10, pady=8)
            self.result_labels[field] = lbl_val

        separator = ttk.Separator(result_grid_frame, orient="horizontal")
        separator.grid(row=5, column=0, columnspan=4, sticky="ew", pady=20)

        big_labels = ["Investment", "Single Profit", "Large Profit", "With Investment"]
        for i, field in enumerate(big_labels):
            lbl_name = ttk.Label(result_grid_frame, text=field + ":", style="BigResult.TLabel")
            lbl_name.grid(row=6 + i, column=0, sticky="e", padx=10, pady=10)

            lbl_val = ttk.Label(result_grid_frame, text="0.00", style="BigResult.TLabel")
            lbl_val.grid(row=6 + i, column=1, sticky="w", padx=10, pady=10)
            self.result_labels[field] = lbl_val

        # Footer with export button
        footer_frame = ttk.Frame(self.root, padding=10, style="TLabel")
        footer_frame.pack(side="bottom", fill="x")

        save_btn = ttk.Button(footer_frame, text="Export to Excel", command=self.export_to_excel)
        save_btn.pack(ipadx=25, ipady=10)

    def create_input_fields(self, parent, fields):
        for field in fields:
            frame = ttk.Frame(parent, style="Card.TFrame")
            frame.pack(fill="x", pady=5)

            lbl = ttk.Label(frame, text=field + ":", style="InputLabel.TLabel", width=22, anchor="e")
            lbl.pack(side="left", padx=(0, 10))

            entry = ttk.Entry(frame, style="TEntry")
            entry.pack(side="left", fill="x", expand=True)
            entry.bind("<KeyRelease>", self.on_input_change)
            entry.bind("<Return>", self.move_focus)
            entry.bind("<FocusIn>", self.start_glow_animation)
            entry.bind("<FocusOut>", self.stop_glow_animation)

            self.entries[field] = entry
            self.fields_order.append(entry)

    def on_input_change(self, event):
        self.update_results()
        self.save_data()

    def save_data(self):
        data = {key: self.entries[key].get() for key in self.entries}
        os.makedirs("dasta", exist_ok=True)
        with open(DATA_FILE, "w") as f:
            json.dump(data, f)

    def load_data(self):
        if os.path.exists(DATA_FILE):
            try:
                with open(DATA_FILE, "r") as f:
                    data = json.load(f)
                for key, val in data.items():
                    if key in self.entries:
                        self.entries[key].delete(0, tk.END)
                        self.entries[key].insert(0, val)
            except Exception as e:
                print(f"Failed to load saved data: {e}")

    def export_to_excel(self):
        try:
            os.makedirs(EXPORT_DIR, exist_ok=True)
            filename = os.path.join(EXPORT_DIR, f"financial_report_{int(time.time())}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Financial Report"

            # Define styles
            bold_font = openpyxl.styles.Font(bold=True)
            border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin'),
            )
            num_fmt = '0.00'

            def apply_header(cell):
                cell.font = bold_font
                cell.border = border

            def apply_cell(cell):
                cell.border = border

            # Write headers
            ws['A1'] = "Field"
            ws['B1'] = "Value"
            apply_header(ws['A1'])
            apply_header(ws['B1'])

            row = 2
            # Inputs section header
            ws[f'A{row}'] = "--- Inputs ---"
            apply_header(ws[f'A{row}'])
            row += 1

            # Inputs
            for key in self.entries:
                ws[f'A{row}'] = key
                ws[f'B{row}'] = self.entries[key].get()
                apply_cell(ws[f'A{row}'])
                apply_cell(ws[f'B{row}'])
                row += 1

            row += 1
            # Results section header
            ws[f'A{row}'] = "--- Results ---"
            apply_header(ws[f'A{row}'])
            row += 1

            # Results
            for key, value in self.current_results.items():
                ws[f'A{row}'] = key
                cell = ws[f'B{row}']
                cell.value = value
                cell.number_format = num_fmt
                apply_cell(ws[f'A{row}'])
                apply_cell(cell)
                row += 1

            # Adjust column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20

            wb.save(filename)
            messagebox.showinfo("Export Successful", f"Data exported to {filename}")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export data:\n{e}")

    def start_glow_animation(self, event):
        widget = event.widget
        if widget not in self.animating:
            self.animating[widget] = True
            self._animate_glow(widget, 0)

    def stop_glow_animation(self, event):
        widget = event.widget
        self.animating[widget] = False
        widget.configure(highlightbackground="#333333", highlightcolor="#333333")

    def _animate_glow(self, widget, step):
        if not self.animating.get(widget, False):
            return
        glow_colors = [
            "#00FFF7", "#00D6CC", "#00B2AA", "#008C88",
            "#006866", "#004644", "#002322", "#000F11",
            "#002322", "#004644", "#006866", "#008C88",
            "#00B2AA", "#00D6CC",
        ]
        color = glow_colors[step % len(glow_colors)]
        widget.configure(highlightbackground=color, highlightcolor=color)

        next_step = (step + 1) % len(glow_colors)
        widget.after(80, lambda: self._animate_glow(widget, next_step))

    def get_value(self, key):
        val = self.entries[key].get().strip()
        try:
            return float(val) if val else 0.0
        except ValueError:
            return 0.0

    def move_focus(self, event):
        widget = event.widget
        try:
            idx = self.fields_order.index(widget)
            if idx + 1 < len(self.fields_order):
                self.fields_order[idx + 1].focus_set()
        except ValueError:
            pass

    def update_results(self):
        sales = self.get_value("Sales")
        cogs = self.get_value("COGS")
        shipping = self.get_value("Shipping")
        warehousing = self.get_value("Warehousing")
        platform_commissions = self.get_value("Platform Commissions")
        marketing = self.get_value("Marketing")
        technology = self.get_value("Technology")
        fixed_cost = self.get_value("Fixed Cost")
        depreciation = self.get_value("Depreciation")
        interest = self.get_value("Interest")
        tax_rate = self.get_value("Tax Rate (%)")
        ad_spend = self.get_value("Ad Spend")
        customers = self.get_value("Customers") or 1

        gross_margin = sales - cogs
        cm1 = gross_margin - shipping - warehousing - platform_commissions
        cm2 = cm1 - (marketing + technology)
        ebitda = cm2 - fixed_cost
        ebit = ebitda - depreciation
        pbt = ebit - interest
        tax = (tax_rate / 100) * pbt
        pat = pbt - tax

        cac = ad_spend / customers
        investment = (cogs * customers) + ad_spend
        single_profit = pat
        large_profit = pat * customers
        with_investment = investment + large_profit

        results = {
            "Gross Margin": gross_margin,
            "Contribution Margin 1": cm1,
            "Contribution Margin 2": cm2,
            "EBITDA": ebitda,
            "EBIT": ebit,
            "PBT": pbt,
            "TAX": tax,
            "PAT": pat,
            "Customer Acquisition Cost": cac,
            "Investment": investment,
            "Single Profit": single_profit,
            "Large Profit": large_profit,
            "With Investment": with_investment
        }

        self.current_results = results

        for key, value in results.items():
            if key in self.result_labels:
                self.result_labels[key].config(text=f"{value:.2f}")

    def schedule_midnight_save(self):
        now = datetime.now()
        next_midnight = datetime.combine(now.date() + timedelta(days=1), datetime.min.time())
        delay = (next_midnight - now).total_seconds()

        def wait_and_save():
            time.sleep(delay)
            self.save_results()
            self.schedule_midnight_save()

        threading.Thread(target=wait_and_save, daemon=True).start()

    def save_results(self):
        self.update_results()
        os.makedirs("dasta", exist_ok=True)
        filename = f"dasta/financial_output_{int(time.time())}.txt"
        with open(filename, "w") as f:
            for key, value in self.current_results.items():
                f.write(f"{key}: {value:.2f}\n")
        print(f"[Auto-Saved] Results saved to {filename}")

if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("Please install openpyxl: pip install openpyxl")
        exit(1)

    root = tk.Tk()
    app = FinancialApp(root)
    root.mainloop()
